"""
HR Views - Employees, Payroll, Payslips with PDF Export
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, make_response
from datetime import date, datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML
from app import api_request, login_required, permission_required
import json

bp = Blueprint('hr', __name__, url_prefix='/hr')


@bp.route('/employees')
@login_required
@permission_required('employees:view')
def list_employees():
    """List all employees"""
    employees, status = api_request('GET', '/hr/employees')
    
    if status != 200:
        employees = []
    
    return render_template('hr/employees.html', title='Employees', employees=employees)


@bp.route('/employees/new', methods=['GET', 'POST'])
@login_required
@permission_required('employees:create')
def new_employee():
    """Create new employee"""
    if request.method == 'GET':
        return render_template('hr/employee_form.html', title='New Employee')
    
    payroll_config = None
    if request.form.get('gross_salary'):
        payroll_config = {
            'gross_salary': request.form.get('gross_salary'),
            'pay_frequency': request.form.get('pay_frequency', 'Monthly'),
            'paye_rate': request.form.get('paye_rate'),
            'pension_employee_rate': request.form.get('pension_employee_rate'),
            'pension_employer_rate': request.form.get('pension_employer_rate')
        }
    
    data = {
        'full_name': request.form.get('full_name'),
        'email': request.form.get('email'),
        'phone_number': request.form.get('phone_number'),
        'address': request.form.get('address'),
        'hire_date': request.form.get('hire_date'),
        'department': request.form.get('department'),
        'position': request.form.get('position'),
        'payroll_config': payroll_config
    }
    
    response, status = api_request('POST', '/hr/employees', data=data)
    
    if status == 200:
        flash('Employee created successfully', 'success')
        return redirect(url_for('hr.view_employee', employee_id=response.get('id')))
    
    error = response.get('detail', 'Failed to create employee') if response else 'Failed'
    flash(error, 'error')
    return render_template('hr/employee_form.html', title='New Employee', error=error)


@bp.route('/employees/<int:employee_id>')
@login_required
@permission_required('employees:view')
def view_employee(employee_id):
    """View employee details"""
    employee, status = api_request('GET', f'/hr/employees/{employee_id}')
    
    if status != 200:
        flash('Employee not found', 'error')
        return redirect(url_for('hr.list_employees'))
    
    return render_template('hr/employee_detail.html',
                          title=employee.get('full_name', 'Employee'),
                          employee=employee)


@bp.route('/employees/<int:employee_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('employees:edit')
def edit_employee(employee_id):
    """Edit employee"""
    if request.method == 'GET':
        employee, status = api_request('GET', f'/hr/employees/{employee_id}')
        if status != 200:
            flash('Employee not found', 'error')
            return redirect(url_for('hr.list_employees'))
        return render_template('hr/employee_form.html', title='Edit Employee', employee=employee)
    
    data = {k: v for k, v in request.form.items() if k != 'csrf_token'}
    
    response, status = api_request('PUT', f'/hr/employees/{employee_id}', data=data)
    
    if status == 200:
        flash('Employee updated', 'success')
        return redirect(url_for('hr.view_employee', employee_id=employee_id))
    
    error = response.get('detail', 'Failed to update employee') if response else 'Failed'
    flash(error, 'error')
    return redirect(url_for('hr.edit_employee', employee_id=employee_id))


@bp.route('/employees/<int:employee_id>/payroll-config', methods=['GET', 'POST'])
@login_required
@permission_required('employees:edit')
def edit_payroll_config(employee_id):
    """Edit employee payroll configuration"""
    employee, status = api_request('GET', f'/hr/employees/{employee_id}')
    if status != 200:
        flash('Employee not found', 'error')
        return redirect(url_for('hr.list_employees'))
    
    if request.method == 'POST':
        data = {
            'gross_salary': request.form.get('gross_salary'),
            'pay_frequency': request.form.get('pay_frequency', 'Monthly'),
            'paye_rate': request.form.get('paye_rate', 0),
            'pension_employee_rate': request.form.get('pension_employee_rate', 0),
            'pension_employer_rate': request.form.get('pension_employer_rate', 0),
            'other_deductions': request.form.get('other_deductions', 0),
            'other_allowances': request.form.get('other_allowances', 0)
        }
        
        # Check if payroll config exists
        existing_config, _ = api_request('GET', f'/hr/employees/{employee_id}/payroll-config')
        
        if existing_config and 'id' in existing_config:
            response, status = api_request('PUT', f'/hr/employees/{employee_id}/payroll-config', data=data)
        else:
            response, status = api_request('POST', f'/hr/employees/{employee_id}/payroll-config', data=data)
        
        if status in [200, 201]:
            flash('Payroll configuration saved', 'success')
            return redirect(url_for('hr.view_employee', employee_id=employee_id))
        
        error = response.get('detail', 'Failed to save payroll configuration') if response else 'Failed'
        flash(error, 'error')
    
    return render_template('hr/payroll_config_form.html', title='Edit Payroll Configuration', employee=employee)


@bp.route('/employees/<int:employee_id>/terminate', methods=['POST'])
@login_required
@permission_required('employees:edit')
def terminate_employee(employee_id):
    """Terminate employee"""
    termination_date = request.form.get('termination_date')
    
    response, status = api_request('POST', f'/hr/employees/{employee_id}/terminate?termination_date={termination_date}')
    
    if status == 200:
        flash('Employee terminated successfully', 'success')
    else:
        error = response.get('detail', 'Failed to terminate employee') if response else 'Failed'
        flash(error, 'error')
    
    return redirect(url_for('hr.view_employee', employee_id=employee_id))


@bp.route('/payroll', methods=['GET', 'POST'])
@login_required
@permission_required('payroll:view')
def run_payroll():
    """Run Payroll"""
    if request.method == 'POST':
        pay_period_start = request.form.get('pay_period_start')
        pay_period_end = request.form.get('pay_period_end')
        
        if pay_period_start and pay_period_end:
            response, status = api_request('POST', '/hr/payroll/run', data={
                'pay_period_start': pay_period_start,
                'pay_period_end': pay_period_end
            })
            
            if status == 200:
                flash(response.get('message', 'Payroll completed successfully'), 'success')
                return redirect(url_for('hr.payslips'))
            else:
                error = response.get('detail', 'Failed to run payroll') if response else 'Failed'
                flash(error, 'error')
    
    employees, _ = api_request('GET', '/hr/employees')
    return render_template('hr/run_payroll.html', title='Run Payroll', employees=employees or [])


@bp.route('/payslips')
@login_required
@permission_required('payroll:view')
def payslips():
    """Payslip History"""
    payslips, status = api_request('GET', '/hr/payslips')
    
    if status != 200:
        payslips = []
    
    return render_template('hr/payslips.html', title='Payslips', payslips=payslips)


@bp.route('/payslips/<int:payslip_id>')
@login_required
@permission_required('payroll:view')
def view_payslip(payslip_id):
    """View payslip details"""
    payslip, status = api_request('GET', f'/hr/payslips/{payslip_id}')
    
    if status != 200:
        flash('Payslip not found', 'error')
        return redirect(url_for('hr.payslips'))
    
    return render_template('hr/payslip_detail.html', title='Payslip', payslip=payslip)


@bp.route('/payslips/<int:payslip_id>/mark-paid', methods=['POST'])
@login_required
@permission_required('payroll:edit')
def mark_payslip_paid(payslip_id):
    """Mark payslip as paid"""
    response, status = api_request('POST', f'/hr/payslips/{payslip_id}/mark-paid')
    
    if status == 200:
        flash('Payslip marked as paid', 'success')
    else:
        error = response.get('detail', 'Failed to mark payslip as paid') if response else 'Failed'
        flash(error, 'error')
    
    return redirect(url_for('hr.view_payslip', payslip_id=payslip_id))


@bp.route('/payslips/<int:payslip_id>/export/pdf')
@login_required
@permission_required('payroll:view')
def payslip_pdf(payslip_id):
    """Export payslip to PDF"""
    payslip, status = api_request('GET', f'/hr/payslips/{payslip_id}')
    
    if status != 200:
        flash('Payslip not found', 'error')
        return redirect(url_for('hr.payslips'))
    
    business = {
        'name': session.get('business_name', 'Company'),
        'branch': session.get('selected_branch_name', 'Main Branch')
    }
    
    html = render_template('reports/pdf/payslip_pdf.html',
                          payslip=payslip,
                          business=business,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=payslip_{payslip.get("payslip_number", payslip_id)}.pdf'
    return response


@bp.route('/payslips/export/excel')
@login_required
@permission_required('payroll:view')
def payslips_excel():
    """Export payslips to Excel"""
    pay_period_start = request.args.get('pay_period_start')
    pay_period_end = request.args.get('pay_period_end')
    
    params = {}
    if pay_period_start:
        params['pay_period_start'] = pay_period_start
    if pay_period_end:
        params['pay_period_end'] = pay_period_end
    
    payslips, _ = api_request('GET', '/hr/payslips', params=params if params else None)
    payslips = payslips or []
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payslips"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Payroll Report"
    ws['A1'].font = header_font
    ws.merge_cells('A1:I1')
    
    if pay_period_start or pay_period_end:
        ws['A2'] = f"Period: {pay_period_start or ''} to {pay_period_end or ''}"
        ws.merge_cells('A2:I2')
    
    # Summary
    total_gross = sum(p.get('gross_salary', 0) or 0 for p in payslips)
    total_deductions = sum(p.get('total_deductions', 0) or 0 for p in payslips)
    total_net = sum(p.get('net_salary', 0) or 0 for p in payslips)
    
    ws['A4'] = "Summary"
    ws['A4'].font = title_font
    ws['A5'] = "Total Employees:"
    ws['B5'] = len(payslips)
    ws['A6'] = "Total Gross:"
    ws['B6'] = total_gross
    ws['B6'].number_format = currency_format
    ws['A7'] = "Total Deductions:"
    ws['B7'] = total_deductions
    ws['B7'].number_format = currency_format
    ws['A8'] = "Total Net:"
    ws['B8'] = total_net
    ws['B8'].number_format = currency_format
    
    # Payslips table
    ws['A10'] = "Payslip Details"
    ws['A10'].font = title_font
    
    headers = ['Payslip #', 'Employee', 'Period Start', 'Period End', 'Basic', 'Allowances', 'Gross', 'Deductions', 'Net Pay', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for row, p in enumerate(payslips, 12):
        employee = p.get('employee', {})
        ws.cell(row=row, column=1, value=p.get('payslip_number')).border = thin_border
        ws.cell(row=row, column=2, value=employee.get('full_name') if isinstance(employee, dict) else str(employee)).border = thin_border
        ws.cell(row=row, column=3, value=p.get('pay_period_start')).border = thin_border
        ws.cell(row=row, column=4, value=p.get('pay_period_end')).border = thin_border
        
        basic_cell = ws.cell(row=row, column=5, value=p.get('basic_salary'))
        basic_cell.number_format = currency_format
        basic_cell.border = thin_border
        
        allow_cell = ws.cell(row=row, column=6, value=p.get('allowances'))
        allow_cell.number_format = currency_format
        allow_cell.border = thin_border
        
        gross_cell = ws.cell(row=row, column=7, value=p.get('gross_salary'))
        gross_cell.number_format = currency_format
        gross_cell.border = thin_border
        
        ded_cell = ws.cell(row=row, column=8, value=p.get('total_deductions'))
        ded_cell.number_format = currency_format
        ded_cell.border = thin_border
        
        net_cell = ws.cell(row=row, column=9, value=p.get('net_salary'))
        net_cell.number_format = currency_format
        net_cell.border = thin_border
        
        ws.cell(row=row, column=10, value=p.get('status', 'pending')).border = thin_border
    
    # Adjust column widths
    col_widths = [12, 25, 12, 12, 12, 12, 12, 12, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=payslips_{date.today()}.xlsx'
    return response


@bp.route('/payroll/summary')
@login_required
@permission_required('payroll:view')
def payroll_summary():
    """Payroll Summary Report"""
    pay_period_start = request.args.get('pay_period_start')
    pay_period_end = request.args.get('pay_period_end')
    
    if not pay_period_start or not pay_period_end:
        pay_period_start = date.today().replace(day=1).isoformat()
        pay_period_end = date.today().isoformat()
    
    summary, status = api_request('GET', f'/hr/payroll/summary?pay_period_start={pay_period_start}&pay_period_end={pay_period_end}')
    
    if status != 200:
        summary = {}
    
    return render_template('hr/payroll_summary.html', 
                          title='Payroll Summary', 
                          summary=summary,
                          pay_period_start=pay_period_start,
                          pay_period_end=pay_period_end)
