"""
Cash Book Routes
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, make_response
from app import api_request, login_required, permission_required
from datetime import date, timedelta, datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML

cashbook_bp = Blueprint('cashbook', __name__, url_prefix='/cashbook')


@cashbook_bp.route('')
@login_required
@permission_required('accounting:view')
def cashbook_list():
    """List cash book entries"""
    # Get query parameters
    start_date = request.args.get('start_date', (date.today() - timedelta(days=30)).isoformat())
    end_date = request.args.get('end_date', date.today().isoformat())
    account_id = request.args.get('account_id', '')
    entry_type = request.args.get('entry_type', '')
    
    params = {
        'start_date': start_date,
        'end_date': end_date
    }
    if account_id:
        params['account_id'] = account_id
    if entry_type:
        params['entry_type'] = entry_type
    
    # Get entries
    entries, status_code = api_request('GET', '/cashbook', params=params)
    
    if status_code != 200:
        entries = []
    
    # Get summary
    summaries, _ = api_request('GET', '/cashbook/summary', params=params)
    
    # Get payment accounts
    payment_accounts, _ = api_request('GET', '/banking/payment-accounts')
    
    # Get currency from session
    currency = session.get('branch_currency', '$')
    
    return render_template('cashbook/list.html',
                          title='Cash Book',
                          cash_book_entries=entries,
                          summaries=summaries or [],
                          payment_accounts=payment_accounts or [],
                          start_date=start_date,
                          end_date=end_date,
                          selected_account=int(account_id) if account_id else None,
                          selected_entry_type=entry_type,
                          total_entries=len(entries),
                          currency=currency)


@cashbook_bp.route('/cash-flow')
@login_required
@permission_required('accounting:view')
def cash_flow():
    """Get cash flow summary"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    summary, status_code = api_request('GET', '/cashbook/cash-flow', params=params if params else None)
    
    if status_code != 200:
        return jsonify({'error': 'Failed to get cash flow summary'}), 500
    
    return jsonify(summary)


@cashbook_bp.route('/account/<int:account_id>')
@login_required
@permission_required('accounting:view')
def account_transactions(account_id):
    """Get transactions for a specific account"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    result, status_code = api_request('GET', f'/cashbook/account/{account_id}', params=params if params else None)
    
    if status_code != 200:
        flash('Account not found', 'error')
        return redirect(url_for('cashbook.cashbook_list'))
    
    return render_template('cashbook/account_transactions.html',
                          title=f"Account Transactions",
                          account=result.get('account', {}),
                          summary=result.get('summary', {}),
                          entries=result.get('entries', []))


@cashbook_bp.route('/entry/<int:entry_id>')
@login_required
@permission_required('accounting:view')
def entry_detail(entry_id):
    """Get a single cash book entry"""
    entry, status_code = api_request('GET', f'/cashbook/entry/{entry_id}')
    
    if status_code != 200:
        flash('Entry not found', 'error')
        return redirect(url_for('cashbook.cashbook_list'))
    
    return render_template('cashbook/entry_detail.html',
                          title=f"Cash Book Entry",
                          entry=entry)


@cashbook_bp.route('/create', methods=['GET', 'POST'])
@login_required
@permission_required('accounting:create')
def create_entry():
    """Create a manual cash book entry"""
    if request.method == 'GET':
        payment_accounts, _ = api_request('GET', '/banking/payment-accounts')
        customers, _ = api_request('GET', '/cashbook/customers-with-balance')
        vendors, _ = api_request('GET', '/cashbook/vendors-with-balance')
        
        # Get currency from session
        currency = session.get('branch_currency', '$')
        
        return render_template('cashbook/form.html',
                              title='New Cash Book Entry',
                              payment_accounts=payment_accounts or [],
                              customers=customers or [],
                              vendors=vendors or [],
                              today=date.today().isoformat(),
                              currency=currency)
    
    # POST - Create entry
    # Check if this is a fund account request
    fund_type = request.form.get('fund_type')
    if fund_type in ['customer', 'vendor']:
        return fund_account()
    
    data = {
        'entry_date': request.form.get('entry_date'),
        'entry_type': request.form.get('entry_type'),
        'account_id': int(request.form.get('account_id')),
        'account_type': request.form.get('account_type', 'cash'),
        'amount': float(request.form.get('amount', 0)),
        'description': request.form.get('description') or None,
        'reference': request.form.get('reference') or None,
        'payee_payer': request.form.get('payee_payer') or None,
    }
    
    response, status_code = api_request('POST', '/cashbook', data=data)
    
    if status_code == 200:
        flash('Cash book entry created successfully', 'success')
        return redirect(url_for('cashbook.cashbook_list'))
    
    error = response.get('detail', 'Failed to create entry') if response else 'Failed'
    flash(error, 'error')
    return redirect(url_for('cashbook.create_entry'))


def fund_account():
    """Fund a customer or vendor account"""
    fund_type = request.form.get('fund_type')
    entity_id = request.form.get('entity_id')
    amount = request.form.get('amount')
    payment_account_id = request.form.get('payment_account_id')
    bank_account_id = request.form.get('bank_account_id')  # Bank account ID if funding from bank
    description = request.form.get('description')
    reference = request.form.get('reference')
    
    if not entity_id or not amount or not payment_account_id:
        flash('Please fill in all required fields', 'error')
        return redirect(url_for('cashbook.create_entry'))
    
    data = {
        'entity_type': fund_type,
        'entity_id': int(entity_id),
        'amount': float(amount),
        'payment_account_id': int(payment_account_id),
        'bank_account_id': int(bank_account_id) if bank_account_id else None,
        'description': description or None,
        'reference': reference or None
    }
    
    response, status_code = api_request('POST', '/cashbook/fund-account', data=data)
    
    if status_code == 200:
        flash(response.get('message', f'{fund_type.title()} account funded successfully'), 'success')
        return redirect(url_for('cashbook.cashbook_list'))
    
    error = response.get('detail', 'Failed to fund account') if response else 'Failed'
    flash(error, 'error')
    return redirect(url_for('cashbook.create_entry'))


@cashbook_bp.route('/entry/<int:entry_id>/delete', methods=['POST'])
@login_required
@permission_required('accounting:delete')
def delete_entry(entry_id):
    """Delete a manual cash book entry"""
    response, status_code = api_request('DELETE', f'/cashbook/entry/{entry_id}')
    
    if status_code == 200:
        flash('Entry deleted successfully', 'success')
    else:
        error = response.get('detail', 'Cannot delete entry') if response else 'Failed'
        flash(error, 'error')
    
    return redirect(url_for('cashbook.cashbook_list'))


@cashbook_bp.route('/reconcile/<int:account_id>', methods=['POST'])
@login_required
@permission_required('accounting:edit')
def reconcile_account(account_id):
    """Reconcile cash book with ledger for an account"""
    as_of_date = request.form.get('as_of_date') or date.today().isoformat()
    
    result, status_code = api_request('POST', f'/cashbook/reconcile/{account_id}', 
                                       params={'as_of_date': as_of_date})
    
    if status_code == 200:
        return jsonify(result)
    
    return jsonify({'error': 'Reconciliation failed'}), 500


@cashbook_bp.route('/export/pdf')
@login_required
@permission_required('accounting:view')
def export_pdf():
    """Export Cash Book to PDF"""
    # Get query parameters
    start_date = request.args.get('start_date', (date.today() - timedelta(days=30)).isoformat())
    end_date = request.args.get('end_date', date.today().isoformat())
    account_id = request.args.get('account_id', '')
    entry_type = request.args.get('entry_type', '')
    
    params = {
        'start_date': start_date,
        'end_date': end_date
    }
    if account_id:
        params['account_id'] = account_id
    if entry_type:
        params['entry_type'] = entry_type
    
    # Get entries and summary
    entries, status_code = api_request('GET', '/cashbook', params=params)
    summaries, _ = api_request('GET', '/cashbook/summary', params=params)
    
    if status_code != 200:
        entries = []
    
    business = {
        'name': session.get('business_name', 'Company'),
        'branch': session.get('selected_branch_name', 'Main Branch')
    }
    
    currency = session.get('branch_currency', '$')
    
    html = render_template('cashbook/pdf/cashbook_pdf.html',
                          entries=entries or [],
                          summaries=summaries or [],
                          business=business,
                          start_date=start_date,
                          end_date=end_date,
                          currency=currency,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=cashbook_{start_date}_to_{end_date}.pdf'
    return response


@cashbook_bp.route('/export/excel')
@login_required
@permission_required('accounting:view')
def export_excel():
    """Export Cash Book to Excel"""
    # Get query parameters
    start_date = request.args.get('start_date', (date.today() - timedelta(days=30)).isoformat())
    end_date = request.args.get('end_date', date.today().isoformat())
    account_id = request.args.get('account_id', '')
    entry_type = request.args.get('entry_type', '')
    
    params = {
        'start_date': start_date,
        'end_date': end_date
    }
    if account_id:
        params['account_id'] = account_id
    if entry_type:
        params['entry_type'] = entry_type
    
    # Get entries and summary
    entries, status_code = api_request('GET', '/cashbook', params=params)
    summaries, _ = api_request('GET', '/cashbook/summary', params=params)
    
    if status_code != 200:
        entries = []
    
    currency = session.get('branch_currency', '$')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Book"
    
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    receipt_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    payment_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    transfer_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"Cash Book - {session.get('business_name', 'Company')}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Period: {start_date} to {end_date}"
    ws.merge_cells('A2:H2')
    
    row = 4
    
    # Account Summaries Section
    if summaries:
        ws.cell(row=row, column=1, value="Account Summaries").font = header_font
        row += 1
        
        summary_headers = ['Account', 'Type', 'Opening Balance', 'Receipts', 'Payments', 'Closing Balance']
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        
        for summary in summaries:
            ws.cell(row=row, column=1, value=summary.get('account_name')).border = thin_border
            ws.cell(row=row, column=2, value=summary.get('account_type')).border = thin_border
            ws.cell(row=row, column=3, value=summary.get('opening_balance')).number_format = currency_format
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4, value=summary.get('total_receipts')).number_format = currency_format
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5, value=summary.get('total_payments')).number_format = currency_format
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=6, value=summary.get('closing_balance')).number_format = currency_format
            ws.cell(row=row, column=6).border = thin_border
            row += 1
        
        row += 2
    
    # Entries Section
    ws.cell(row=row, column=1, value="Cash Book Entries").font = header_font
    row += 1
    
    entry_headers = ['Entry #', 'Date', 'Type', 'Account', 'Description', 'Payee/Payer', 'Amount', 'Source']
    for col, header in enumerate(entry_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    total_receipts = 0
    total_payments = 0
    
    for entry in (entries or []):
        entry_type_val = entry.get('entry_type', '')
        
        ws.cell(row=row, column=1, value=entry.get('entry_number')).border = thin_border
        ws.cell(row=row, column=2, value=entry.get('entry_date')).border = thin_border
        
        type_cell = ws.cell(row=row, column=3, value=entry_type_val.title())
        type_cell.border = thin_border
        if entry_type_val == 'receipt':
            type_cell.fill = receipt_fill
        elif entry_type_val == 'payment':
            type_cell.fill = payment_fill
        else:
            type_cell.fill = transfer_fill
        
        ws.cell(row=row, column=4, value=entry.get('account_name', '-')).border = thin_border
        ws.cell(row=row, column=5, value=entry.get('description', '-')).border = thin_border
        ws.cell(row=row, column=6, value=entry.get('payee_payer', '-')).border = thin_border
        
        amount = entry.get('amount', 0)
        amount_cell = ws.cell(row=row, column=7, value=amount)
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
        
        source = entry.get('source_type', 'Manual')
        ws.cell(row=row, column=8, value=source.replace('_', ' ').title() if source else 'Manual').border = thin_border
        
        # Track totals
        if entry_type_val == 'receipt':
            total_receipts += amount
        elif entry_type_val == 'payment':
            total_payments += amount
        
        row += 1
    
    # Totals row
    row += 1
    ws.cell(row=row, column=5, value="Total Receipts:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_receipts).number_format = currency_format
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=6).fill = receipt_fill
    row += 1
    
    ws.cell(row=row, column=5, value="Total Payments:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_payments).number_format = currency_format
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=6).fill = payment_fill
    row += 1
    
    ws.cell(row=row, column=5, value="Net Cash Flow:").font = Font(bold=True, size=12)
    net_flow = total_receipts - total_payments
    net_cell = ws.cell(row=row, column=6, value=net_flow)
    net_cell.number_format = currency_format
    net_cell.font = Font(bold=True, size=12)
    net_cell.fill = total_fill
    row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=cashbook_{start_date}_to_{end_date}.xlsx'
    return response
