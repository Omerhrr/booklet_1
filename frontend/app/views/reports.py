"""
Reports Views with PDF and Excel Export
"""
from flask import Blueprint, render_template, request, Response, session, make_response
from datetime import date, datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML, CSS
from app import api_request, login_required, permission_required

bp = Blueprint('reports', __name__, url_prefix='/reports')

# Get business info for reports
def get_business_info():
    """Get business info for report headers"""
    return {
        'name': session.get('business_name', 'Company'),
        'branch': session.get('selected_branch_name', 'Main Branch')
    }


# ==================== REPORTS INDEX ====================

@bp.route('')
@login_required
@permission_required('reports:view')
def index():
    """Reports index"""
    return render_template('reports/index.html', title='Reports')


# ==================== SALES REPORT ====================

@bp.route('/sales')
@login_required
@permission_required('reports:view')
def sales_report():
    """Sales Report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, status = api_request('GET', '/reports/sales', params=params if params else None)
    
    if status != 200:
        report = {}
    
    return render_template('reports/sales_report.html', 
                          title='Sales Report', 
                          report=report,
                          start_date=start_date,
                          end_date=end_date)


@bp.route('/sales/export/pdf')
@login_required
@permission_required('reports:view')
def sales_report_pdf():
    """Export Sales Report to PDF"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, _ = api_request('GET', '/reports/sales', params=params if params else None)
    business = get_business_info()
    
    html = render_template('reports/pdf/sales_report_pdf.html',
                          report=report or {},
                          business=business,
                          start_date=start_date,
                          end_date=end_date,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{date.today()}.pdf'
    return response


@bp.route('/sales/export/excel')
@login_required
@permission_required('reports:view')
def sales_report_excel():
    """Export Sales Report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, _ = api_request('GET', '/reports/sales', params=params if params else None)
    report = report or {}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Sales Report"
    ws['A1'].font = header_font
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Period: {report.get('start_date', '')} to {report.get('end_date', '')}"
    ws.merge_cells('A2:E2')
    
    # Summary
    ws['A4'] = "Summary"
    ws['A4'].font = title_font
    
    ws['A5'] = "Total Sales:"
    ws['B5'] = report.get('total_sales', 0)
    ws['B5'].number_format = currency_format
    
    ws['A6'] = "Total Invoices:"
    ws['B6'] = report.get('total_invoices', 0)
    
    ws['A7'] = "Collected:"
    ws['B7'] = report.get('collected', 0)
    ws['B7'].number_format = currency_format
    
    ws['A8'] = "Outstanding:"
    ws['B8'] = report.get('outstanding', 0)
    ws['B8'].number_format = currency_format
    
    # Invoices table
    ws['A10'] = "Invoice Details"
    ws['A10'].font = title_font
    
    headers = ['Invoice #', 'Date', 'Customer', 'Amount', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for row, inv in enumerate(report.get('invoices', []), 12):
        ws.cell(row=row, column=1, value=inv.get('invoice_number')).border = thin_border
        ws.cell(row=row, column=2, value=inv.get('invoice_date')).border = thin_border
        ws.cell(row=row, column=3, value=inv.get('customer_name')).border = thin_border
        amount_cell = ws.cell(row=row, column=4, value=inv.get('total_amount'))
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
        ws.cell(row=row, column=5, value=inv.get('status')).border = thin_border
    
    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{date.today()}.xlsx'
    return response


# ==================== PURCHASES REPORT ====================

@bp.route('/purchases')
@login_required
@permission_required('reports:view')
def purchases_report():
    """Purchases Report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, status = api_request('GET', '/reports/purchases', params=params if params else None)
    
    if status != 200:
        report = {}
    
    return render_template('reports/purchase_report.html', 
                          title='Purchases Report', 
                          report=report,
                          start_date=start_date,
                          end_date=end_date)


@bp.route('/purchases/export/pdf')
@login_required
@permission_required('reports:view')
def purchases_report_pdf():
    """Export Purchases Report to PDF"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, _ = api_request('GET', '/reports/purchases', params=params if params else None)
    business = get_business_info()
    
    html = render_template('reports/pdf/purchases_report_pdf.html',
                          report=report or {},
                          business=business,
                          start_date=start_date,
                          end_date=end_date,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=purchases_report_{date.today()}.pdf'
    return response


@bp.route('/purchases/export/excel')
@login_required
@permission_required('reports:view')
def purchases_report_excel():
    """Export Purchases Report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, _ = api_request('GET', '/reports/purchases', params=params if params else None)
    report = report or {}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases Report"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Purchases Report"
    ws['A1'].font = header_font
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Period: {report.get('start_date', '')} to {report.get('end_date', '')}"
    ws.merge_cells('A2:E2')
    
    # Summary
    ws['A4'] = "Summary"
    ws['A4'].font = title_font
    
    ws['A5'] = "Total Purchases:"
    ws['B5'] = report.get('total_purchases', 0)
    ws['B5'].number_format = currency_format
    
    ws['A6'] = "Total Bills:"
    ws['B6'] = report.get('total_bills', 0)
    
    ws['A7'] = "Paid:"
    ws['B7'] = report.get('paid', 0)
    ws['B7'].number_format = currency_format
    
    ws['A8'] = "Outstanding:"
    ws['B8'] = report.get('outstanding', 0)
    ws['B8'].number_format = currency_format
    
    # Bills table
    ws['A10'] = "Bill Details"
    ws['A10'].font = title_font
    
    headers = ['Bill #', 'Date', 'Vendor', 'Amount', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for row, bill in enumerate(report.get('bills', []), 12):
        ws.cell(row=row, column=1, value=bill.get('bill_number')).border = thin_border
        ws.cell(row=row, column=2, value=bill.get('bill_date')).border = thin_border
        ws.cell(row=row, column=3, value=bill.get('vendor_name')).border = thin_border
        amount_cell = ws.cell(row=row, column=4, value=bill.get('total_amount'))
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
        ws.cell(row=row, column=5, value=bill.get('status')).border = thin_border
    
    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=purchases_report_{date.today()}.xlsx'
    return response


# ==================== EXPENSES REPORT ====================

@bp.route('/expenses')
@login_required
@permission_required('reports:view')
def expenses_report():
    """Expenses Report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, status = api_request('GET', '/reports/expenses', params=params if params else None)
    
    if status != 200:
        report = {}
    
    return render_template('reports/expense_report.html', 
                          title='Expenses Report', 
                          report=report,
                          start_date=start_date,
                          end_date=end_date)


@bp.route('/expenses/export/pdf')
@login_required
@permission_required('reports:view')
def expenses_report_pdf():
    """Export Expenses Report to PDF"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, _ = api_request('GET', '/reports/expenses', params=params if params else None)
    business = get_business_info()
    
    html = render_template('reports/pdf/expenses_report_pdf.html',
                          report=report or {},
                          business=business,
                          start_date=start_date,
                          end_date=end_date,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=expenses_report_{date.today()}.pdf'
    return response


@bp.route('/expenses/export/excel')
@login_required
@permission_required('reports:view')
def expenses_report_excel():
    """Export Expenses Report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, _ = api_request('GET', '/reports/expenses', params=params if params else None)
    report = report or {}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses Report"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Expenses Report"
    ws['A1'].font = header_font
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Period: {report.get('start_date', '')} to {report.get('end_date', '')}"
    ws.merge_cells('A2:E2')
    
    # Summary
    ws['A4'] = "Summary"
    ws['A4'].font = title_font
    
    ws['A5'] = "Total Expenses:"
    ws['B5'] = report.get('total_expenses', 0)
    ws['B5'].number_format = currency_format
    
    ws['A6'] = "Total VAT:"
    ws['B6'] = report.get('total_vat', 0)
    ws['B6'].number_format = currency_format
    
    ws['A7'] = "Expense Count:"
    ws['B7'] = report.get('expense_count', 0)
    
    # By Category
    ws['A9'] = "By Category"
    ws['A9'].font = title_font
    
    headers = ['Category', 'Total', 'Count']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    for row, cat in enumerate(report.get('by_category', []), 11):
        ws.cell(row=row, column=1, value=cat.get('category')).border = thin_border
        total_cell = ws.cell(row=row, column=2, value=cat.get('total'))
        total_cell.number_format = currency_format
        total_cell.border = thin_border
        ws.cell(row=row, column=3, value=cat.get('count')).border = thin_border
    
    # Adjust column widths
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=expenses_report_{date.today()}.xlsx'
    return response


# ==================== INVENTORY REPORT ====================

@bp.route('/inventory')
@login_required
@permission_required('reports:view')
def inventory_report():
    """Inventory Report"""
    report, status = api_request('GET', '/reports/inventory')
    
    if status != 200:
        report = {}
    
    return render_template('reports/inventory_report.html', 
                          title='Inventory Report', 
                          report=report)


@bp.route('/inventory/export/pdf')
@login_required
@permission_required('reports:view')
def inventory_report_pdf():
    """Export Inventory Report to PDF"""
    report, _ = api_request('GET', '/reports/inventory')
    business = get_business_info()
    
    html = render_template('reports/pdf/inventory_report_pdf.html',
                          report=report or {},
                          business=business,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=inventory_report_{date.today()}.pdf'
    return response


@bp.route('/inventory/export/excel')
@login_required
@permission_required('reports:view')
def inventory_report_excel():
    """Export Inventory Report to Excel"""
    report, _ = api_request('GET', '/reports/inventory')
    report = report or {}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    number_format = '#,##0.00'
    header_fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Inventory Report"
    ws['A1'].font = header_font
    ws.merge_cells('A1:G1')
    
    # Summary
    ws['A3'] = "Summary"
    ws['A3'].font = title_font
    
    ws['A4'] = "Total Products:"
    ws['B4'] = report.get('total_products', 0)
    
    ws['A5'] = "Total Value:"
    ws['B5'] = report.get('total_value', 0)
    ws['B5'].number_format = currency_format
    
    ws['A6'] = "Low Stock:"
    ws['B6'] = report.get('low_stock_count', 0)
    
    ws['A7'] = "Out of Stock:"
    ws['B7'] = report.get('out_of_stock_count', 0)
    
    # Products table
    ws['A9'] = "Product Details"
    ws['A9'].font = title_font
    
    headers = ['SKU', 'Name', 'Category', 'Quantity', 'Unit', 'Purchase Price', 'Value']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    for row, prod in enumerate(report.get('products', []), 11):
        ws.cell(row=row, column=1, value=prod.get('sku')).border = thin_border
        ws.cell(row=row, column=2, value=prod.get('name')).border = thin_border
        ws.cell(row=row, column=3, value=prod.get('category')).border = thin_border
        qty_cell = ws.cell(row=row, column=4, value=prod.get('quantity'))
        qty_cell.number_format = number_format
        qty_cell.border = thin_border
        ws.cell(row=row, column=5, value=prod.get('unit')).border = thin_border
        price_cell = ws.cell(row=row, column=6, value=prod.get('purchase_price'))
        price_cell.number_format = currency_format
        price_cell.border = thin_border
        value_cell = ws.cell(row=row, column=7, value=prod.get('value'))
        value_cell.number_format = currency_format
        value_cell.border = thin_border
    
    # Adjust column widths
    col_widths = [12, 25, 15, 10, 8, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=inventory_report_{date.today()}.xlsx'
    return response


# ==================== AGING REPORT ====================

@bp.route('/aging')
@login_required
@permission_required('reports:view')
def aging_report():
    """Aging Report"""
    aging, status = api_request('GET', '/dashboard/aging')
    
    if status != 200 or not aging:
        aging = {
            'receivables': {
                'current': 0, 'days_30': 0, 'days_60': 0, 'days_90': 0, 'over_90': 0, 'total': 0
            }, 
            'payables': {
                'current': 0, 'days_30': 0, 'days_60': 0, 'days_90': 0, 'over_90': 0, 'total': 0
            }
        }
    
    return render_template('reports/aging_report.html', 
                          title='Aging Report', 
                          aging=aging)


@bp.route('/aging/export/pdf')
@login_required
@permission_required('reports:view')
def aging_report_pdf():
    """Export Aging Report to PDF"""
    aging, _ = api_request('GET', '/dashboard/aging')
    business = get_business_info()
    
    html = render_template('reports/pdf/aging_report_pdf.html',
                          aging=aging or {},
                          business=business,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=aging_report_{date.today()}.pdf'
    return response


@bp.route('/aging/export/excel')
@login_required
@permission_required('reports:view')
def aging_report_excel():
    """Export Aging Report to Excel"""
    aging, _ = api_request('GET', '/dashboard/aging')
    aging = aging or {}
    
    wb = openpyxl.Workbook()
    
    # Receivables sheet
    ws1 = wb.active
    ws1.title = "Receivables Aging"
    
    header_font = Font(bold=True, size=14)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws1['A1'] = "Receivables Aging Report"
    ws1['A1'].font = header_font
    ws1.merge_cells('A1:F1')
    
    receivables = aging.get('receivables', {})
    headers = ['Period', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    periods = ['current', 'days_30', 'days_60', 'days_90', 'over_90']
    labels = ['Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days']
    
    for row, (period, label) in enumerate(zip(periods, labels), 4):
        ws1.cell(row=row, column=1, value=label).border = thin_border
        amount_cell = ws1.cell(row=row, column=2, value=receivables.get(period, 0))
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
    
    # Payables sheet
    ws2 = wb.create_sheet("Payables Aging")
    
    ws2['A1'] = "Payables Aging Report"
    ws2['A1'].font = header_font
    ws2.merge_cells('A1:F1')
    
    payables = aging.get('payables', {})
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    for row, (period, label) in enumerate(zip(periods, labels), 4):
        ws2.cell(row=row, column=1, value=label).border = thin_border
        amount_cell = ws2.cell(row=row, column=2, value=payables.get(period, 0))
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
    
    # Adjust column widths
    for ws in [ws1, ws2]:
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=aging_report_{date.today()}.xlsx'
    return response


# ==================== TRIAL BALANCE ====================

@bp.route('/trial-balance')
@login_required
@permission_required('reports:view')
def trial_balance():
    """Trial Balance"""
    as_of_date = request.args.get('as_of_date')
    
    params = {}
    if as_of_date:
        params['as_of_date'] = as_of_date
    
    report, status = api_request('GET', '/reports/trial-balance', params=params if params else None)
    
    if status != 200:
        report = {}
    
    return render_template('reports/trial_balance.html', 
                          title='Trial Balance', 
                          report=report,
                          as_of_date=as_of_date)


@bp.route('/trial-balance/export/pdf')
@login_required
@permission_required('reports:view')
def trial_balance_pdf():
    """Export Trial Balance to PDF"""
    as_of_date = request.args.get('as_of_date')
    
    params = {}
    if as_of_date:
        params['as_of_date'] = as_of_date
    
    report, _ = api_request('GET', '/reports/trial-balance', params=params if params else None)
    business = get_business_info()
    
    html = render_template('reports/pdf/trial_balance_pdf.html',
                          report=report or {},
                          business=business,
                          as_of_date=as_of_date,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=trial_balance_{date.today()}.pdf'
    return response


@bp.route('/trial-balance/export/excel')
@login_required
@permission_required('reports:view')
def trial_balance_excel():
    """Export Trial Balance to Excel"""
    as_of_date = request.args.get('as_of_date')
    
    params = {}
    if as_of_date:
        params['as_of_date'] = as_of_date
    
    report, _ = api_request('GET', '/reports/trial-balance', params=params if params else None)
    report = report or {}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    
    header_font = Font(bold=True, size=14)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws['A1'] = "Trial Balance"
    ws['A1'].font = header_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"As of: {report.get('as_of_date', date.today().isoformat())}"
    ws.merge_cells('A2:D2')
    
    headers = ['Code', 'Account', 'Debit', 'Credit']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    for row, acc in enumerate(report.get('accounts', []), 5):
        ws.cell(row=row, column=1, value=acc.get('account_code')).border = thin_border
        ws.cell(row=row, column=2, value=acc.get('account_name')).border = thin_border
        debit_cell = ws.cell(row=row, column=3, value=acc.get('debit') if acc.get('debit') else None)
        debit_cell.number_format = currency_format
        debit_cell.border = thin_border
        credit_cell = ws.cell(row=row, column=4, value=acc.get('credit') if acc.get('credit') else None)
        credit_cell.number_format = currency_format
        credit_cell.border = thin_border
    
    # Totals row
    total_row = len(report.get('accounts', [])) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=report.get('total_debit')).number_format = currency_format
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=report.get('total_credit')).number_format = currency_format
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    
    # Adjust column widths
    col_widths = [12, 30, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=trial_balance_{date.today()}.xlsx'
    return response


# ==================== VAT REPORT ====================

@bp.route('/vat')
@login_required
@permission_required('reports:view')
def vat_report():
    """VAT Report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, status = api_request('GET', '/reports/vat', params=params if params else None)
    
    if status != 200:
        report = {}
    
    return render_template('reports/vat_report.html', 
                          title='VAT Report', 
                          report=report,
                          start_date=start_date,
                          end_date=end_date)


@bp.route('/vat/export/pdf')
@login_required
@permission_required('reports:view')
def vat_report_pdf():
    """Export VAT Report to PDF"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, _ = api_request('GET', '/reports/vat', params=params if params else None)
    business = get_business_info()
    
    html = render_template('reports/pdf/vat_report_pdf.html',
                          report=report or {},
                          business=business,
                          start_date=start_date,
                          end_date=end_date,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=vat_report_{date.today()}.pdf'
    return response


@bp.route('/vat/export/excel')
@login_required
@permission_required('reports:view')
def vat_report_excel():
    """Export VAT Report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    params = {}
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    report, _ = api_request('GET', '/reports/vat', params=params if params else None)
    report = report or {}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VAT Report"
    
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws['A1'] = "VAT Report"
    ws['A1'].font = header_font
    ws.merge_cells('A1:C1')
    
    ws['A2'] = f"Period: {report.get('start_date', '')} to {report.get('end_date', '')}"
    ws.merge_cells('A2:C2')
    
    # Summary
    ws['A4'] = "Summary"
    ws['A4'].font = title_font
    
    ws['A5'] = "VAT Collected (Sales):"
    ws['B5'] = report.get('vat_collected', 0)
    ws['B5'].number_format = currency_format
    
    ws['A6'] = "VAT Paid (Purchases):"
    ws['B6'] = report.get('vat_paid_purchases', 0)
    ws['B6'].number_format = currency_format
    
    ws['A7'] = "VAT Paid (Expenses):"
    ws['B7'] = report.get('vat_paid_expenses', 0)
    ws['B7'].number_format = currency_format
    
    ws['A8'] = "Total VAT Paid:"
    ws['B8'] = report.get('total_vat_paid', 0)
    ws['B8'].number_format = currency_format
    
    ws['A10'] = "Net VAT:"
    ws['B10'] = report.get('net_vat', 0)
    ws['B10'].number_format = currency_format
    ws['B10'].font = Font(bold=True)
    
    ws['A11'] = "VAT Payable:"
    ws['B11'] = report.get('vat_payable', 0)
    ws['B11'].number_format = currency_format
    
    ws['A12'] = "VAT Receivable:"
    ws['B12'] = report.get('vat_receivable', 0)
    ws['B12'].number_format = currency_format
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=vat_report_{date.today()}.xlsx'
    return response
