"""
Accounting Views - Chart of Accounts, Journal, Financial Reports with PDF/Excel Export
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, make_response
from datetime import date, datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML
from app import api_request, login_required, permission_required
import json

bp = Blueprint('accounting', __name__, url_prefix='/accounting')


@bp.route('/chart-of-accounts')
@login_required
@permission_required('accounts:view')
def chart_of_accounts():
    """Chart of Accounts"""
    accounts, status = api_request('GET', '/accounting/accounts')
    
    if status != 200:
        accounts = []
    
    return render_template('accounting/chart_of_accounts.html', title='Chart of Accounts', accounts=accounts)


@bp.route('/accounts/<int:account_id>')
@login_required
@permission_required('accounts:view')
def view_account(account_id):
    """View account details"""
    account, status = api_request('GET', f'/accounting/accounts/{account_id}')
    
    if status != 200:
        flash('Account not found', 'error')
        return redirect(url_for('accounting.chart_of_accounts'))
    
    # Get balance
    balance_data, balance_status = api_request('GET', f'/accounting/accounts/{account_id}/balance')
    
    if balance_status != 200:
        print(f"Warning: Balance API returned status {balance_status}")
        balance_data = {}
    
    # Extract account info from balance response (which includes both account and balance)
    account_info = balance_data.get('account', account) if balance_data else account
    
    # Get ledger entries for this account
    ledger_data, ledger_status = api_request('GET', f'/accounting/accounts/{account_id}/ledger')
    ledger_entries = ledger_data.get('entries', []) if ledger_status == 200 and ledger_data else []
    
    return render_template('accounting/account_detail.html', 
                          title=account_info.get('name', 'Account') if isinstance(account_info, dict) else getattr(account_info, 'name', 'Account'), 
                          account=account_info, 
                          balance=balance_data or {},
                          ledger_entries=ledger_entries)


@bp.route('/accounts/new', methods=['POST'])
@login_required
@permission_required('accounts:create')
def create_account():
    """Create new account"""
    data = {
        'name': request.form.get('name'),
        'code': request.form.get('code'),
        'type': request.form.get('type'),
        'description': request.form.get('description'),
        'parent_id': request.form.get('parent_id')
    }
    
    response, status = api_request('POST', '/accounting/accounts', data=data)
    
    if status == 200:
        flash('Account created', 'success')
        return redirect(url_for('accounting.chart_of_accounts'))
    
    error = response.get('detail', 'Failed to create account') if response else 'Failed'
    flash(error, 'error')
    return redirect(url_for('accounting.chart_of_accounts'))


@bp.route('/journal')
@login_required
@permission_required('journal:view')
def journal():
    """Journal entries list"""
    journals, status = api_request('GET', '/accounting/journal')
    
    if status != 200:
        journals = []
    
    return render_template('accounting/journal_list.html', title='Journal Entries', journals=journals)


@bp.route('/journal/<int:journal_id>')
@login_required
@permission_required('journal:view')
def view_journal(journal_id):
    """View journal entry details"""
    journal, status = api_request('GET', f'/accounting/journal/{journal_id}')
    
    if status != 200:
        flash('Journal entry not found', 'error')
        return redirect(url_for('accounting.journal'))
    
    return render_template('accounting/journal_detail.html', title=f'Journal {journal.get("voucher_number", "")}', journal=journal)


@bp.route('/journal/new', methods=['GET', 'POST'])
@login_required
@permission_required('journal:create')
def new_journal():
    """Create journal entry"""
    if request.method == 'GET':
        accounts, _ = api_request('GET', '/accounting/accounts')
        return render_template('accounting/journal_form.html', title='New Journal Entry', accounts=accounts or [])
    
    lines_json = request.form.get('lines_json', '[]')
    
    data = {
        'transaction_date': request.form.get('transaction_date'),
        'description': request.form.get('description'),
        'reference': request.form.get('reference'),
        'lines': json.loads(lines_json)
    }
    
    response, status = api_request('POST', '/accounting/journal', data=data)
    
    if status == 200:
        flash('Journal entry created', 'success')
        return redirect(url_for('accounting.journal'))
    
    error = response.get('detail', 'Failed to create journal entry') if response else 'Failed'
    flash(error, 'error')
    return redirect(url_for('accounting.new_journal'))


@bp.route('/journal/<int:journal_id>/post', methods=['POST'])
@login_required
@permission_required('journal:create')
def post_journal(journal_id):
    """Post journal entry"""
    response, status = api_request('POST', f'/accounting/journal-vouchers/{journal_id}/post')
    
    if status == 200:
        flash('Journal entry posted', 'success')
    else:
        error = response.get('detail', 'Failed to post journal entry') if response else 'Failed'
        flash(error, 'error')
    
    return redirect(url_for('accounting.view_journal', journal_id=journal_id))


@bp.route('/ledger')
@login_required
@permission_required('accounts:view')
def general_ledger():
    """General Ledger"""
    # Get filter parameters
    account_id = request.args.get('account_id', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build API URL with filters
    api_url = '/accounting/reports/general-ledger?'
    if account_id:
        api_url += f'account_id={account_id}&'
    if start_date:
        api_url += f'start_date={start_date}&'
    if end_date:
        api_url += f'end_date={end_date}&'
    
    # Fetch data
    ledger_data, status = api_request('GET', api_url)
    accounts, _ = api_request('GET', '/accounting/accounts')
    
    entries = []
    summary = {}
    
    if status == 200 and ledger_data:
        entries = ledger_data.get('entries', []) if isinstance(ledger_data, dict) else ledger_data
        summary = ledger_data.get('summary', {}) if isinstance(ledger_data, dict) else {}
    
    return render_template('accounting/general_ledger.html', 
                          title='General Ledger', 
                          entries=entries,
                          summary=summary,
                          accounts=accounts or [],
                          selected_account=account_id,
                          start_date=start_date,
                          end_date=end_date)


@bp.route('/ledger/export/pdf')
@login_required
@permission_required('accounts:view')
def general_ledger_pdf():
    """Export General Ledger to PDF"""
    # Get filter parameters
    account_id = request.args.get('account_id', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build API URL with filters
    api_url = '/accounting/reports/general-ledger?'
    if account_id:
        api_url += f'account_id={account_id}&'
    if start_date:
        api_url += f'start_date={start_date}&'
    if end_date:
        api_url += f'end_date={end_date}&'
    
    # Fetch data
    ledger_data, status = api_request('GET', api_url)
    
    entries = []
    summary = {}
    
    if status == 200 and ledger_data:
        entries = ledger_data.get('entries', []) if isinstance(ledger_data, dict) else ledger_data
        summary = ledger_data.get('summary', {}) if isinstance(ledger_data, dict) else {}
    
    # Get selected account name if filtering by account
    selected_account_name = "All Accounts"
    if account_id:
        accounts, _ = api_request('GET', '/accounting/accounts')
        for acc in (accounts or []):
            if str(acc.get('id')) == str(account_id):
                selected_account_name = f"{acc.get('code')} - {acc.get('name')}"
                break
    
    business = {
        'name': session.get('business_name', 'Company'),
        'branch': session.get('selected_branch_name', 'Main Branch')
    }
    
    html = render_template('accounting/pdf/general_ledger_pdf.html',
                          entries=entries,
                          summary=summary,
                          business=business,
                          selected_account_name=selected_account_name,
                          start_date=start_date,
                          end_date=end_date,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=general_ledger_{start_date or "all"}_to_{end_date or "now"}.pdf'
    return response


@bp.route('/ledger/export/excel')
@login_required
@permission_required('accounts:view')
def general_ledger_excel():
    """Export General Ledger to Excel"""
    # Get filter parameters
    account_id = request.args.get('account_id', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build API URL with filters
    api_url = '/accounting/reports/general-ledger?'
    if account_id:
        api_url += f'account_id={account_id}&'
    if start_date:
        api_url += f'start_date={start_date}&'
    if end_date:
        api_url += f'end_date={end_date}&'
    
    # Fetch data
    ledger_data, status = api_request('GET', api_url)
    
    entries = []
    summary = {}
    
    if status == 200 and ledger_data:
        entries = ledger_data.get('entries', []) if isinstance(ledger_data, dict) else ledger_data
        summary = ledger_data.get('summary', {}) if isinstance(ledger_data, dict) else {}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General Ledger"
    
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    debit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    credit_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    opening_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"General Ledger - {session.get('business_name', 'Company')}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    ws['A2'] = f"Period: {start_date or 'All'} to {end_date or 'Present'}"
    ws.merge_cells('A2:G2')
    
    row = 4
    
    # Summary section if available
    if summary and summary.get('accounts_summary'):
        ws.cell(row=row, column=1, value="Accounts Summary").font = header_font
        row += 1
        
        summary_headers = ['Code', 'Account', 'Type', 'Opening', 'Debits', 'Credits', 'Closing']
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        
        for acc in summary.get('accounts_summary', []):
            ws.cell(row=row, column=1, value=acc.get('account_code')).border = thin_border
            ws.cell(row=row, column=2, value=acc.get('account_name')).border = thin_border
            ws.cell(row=row, column=3, value=acc.get('account_type')).border = thin_border
            ws.cell(row=row, column=4, value=acc.get('opening_balance')).number_format = currency_format
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5, value=acc.get('total_debit')).number_format = currency_format
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=6, value=acc.get('total_credit')).number_format = currency_format
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=7, value=acc.get('closing_balance')).number_format = currency_format
            ws.cell(row=row, column=7).border = thin_border
            row += 1
        
        row += 2
    
    # Entries section
    ws.cell(row=row, column=1, value="Ledger Entries").font = header_font
    row += 1
    
    entry_headers = ['Date', 'Account', 'Description', 'Reference', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(entry_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    for entry in entries:
        is_opening = entry.get('is_opening', False)
        
        date_cell = ws.cell(row=row, column=1, value=entry.get('transaction_date'))
        date_cell.border = thin_border
        if is_opening:
            date_cell.fill = opening_fill
        
        account_cell = ws.cell(row=row, column=2, value=f"{entry.get('account_code', '')} {entry.get('account_name', '')}")
        account_cell.border = thin_border
        if is_opening:
            account_cell.fill = opening_fill
        
        desc_cell = ws.cell(row=row, column=3, value=entry.get('description', '-'))
        desc_cell.border = thin_border
        if is_opening:
            desc_cell.fill = opening_fill
        
        ref_cell = ws.cell(row=row, column=4, value=entry.get('reference', '-'))
        ref_cell.border = thin_border
        if is_opening:
            ref_cell.fill = opening_fill
        
        debit = entry.get('debit', 0)
        debit_cell = ws.cell(row=row, column=5, value=debit if debit > 0 else None)
        debit_cell.number_format = currency_format
        debit_cell.border = thin_border
        if debit > 0:
            debit_cell.fill = debit_fill
        elif is_opening:
            debit_cell.fill = opening_fill
        
        credit = entry.get('credit', 0)
        credit_cell = ws.cell(row=row, column=6, value=credit if credit > 0 else None)
        credit_cell.number_format = currency_format
        credit_cell.border = thin_border
        if credit > 0:
            credit_cell.fill = credit_fill
        elif is_opening:
            credit_cell.fill = opening_fill
        
        balance = entry.get('balance', 0)
        balance_cell = ws.cell(row=row, column=7, value=balance)
        balance_cell.number_format = currency_format
        balance_cell.border = thin_border
        balance_cell.font = Font(bold=True)
        
        row += 1
    
    # Totals row
    row += 1
    ws.cell(row=row, column=4, value="Totals:").font = Font(bold=True)
    ws.cell(row=row, column=5, value=summary.get('total_debit')).number_format = currency_format
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).fill = total_fill
    ws.cell(row=row, column=6, value=summary.get('total_credit')).number_format = currency_format
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=6).fill = total_fill
    ws.cell(row=row, column=7, value=summary.get('net_movement')).number_format = currency_format
    ws.cell(row=row, column=7).font = Font(bold=True)
    ws.cell(row=row, column=7).fill = total_fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=general_ledger_{start_date or "all"}_to_{end_date or "now"}.xlsx'
    return response


@bp.route('/balance-sheet')
@login_required
@permission_required('reports:view')
def balance_sheet():
    """Balance Sheet Report - Statement of Financial Position"""
    as_of_date = request.args.get('as_of_date', '')
    
    report = {
        'non_current_assets': {'fixed_assets_cost': 0, 'accumulated_depreciation': 0, 'net_book_value': 0, 'other_non_current': 0, 'total': 0},
        'current_assets': {'inventory': 0, 'accounts_receivable': 0, 'vat_receivable': 0, 'cash_and_bank': 0, 'other_current_assets': 0, 'vendor_advances': 0, 'total': 0},
        'total_assets': 0,
        'liabilities': {'accounts_payable': 0, 'payroll_liabilities': 0, 'paye_payable': 0, 'pension_payable': 0, 'vat_payable': 0, 'customer_advances': 0, 'other_liabilities': 0, 'total': 0},
        'equity': {'owners_equity': 0, 'retained_earnings': 0, 'opening_balance_equity': 0, 'current_period_earnings': 0, 'total': 0},
        'total_liabilities': 0,
        'total_equity': 0,
        'total_equity_and_liabilities': 0,
        'as_of_date': as_of_date
    }
    
    if as_of_date:
        api_url = f'/accounting/reports/balance-sheet?as_of_date={as_of_date}'
        data, status = api_request('GET', api_url)
        
        if status == 200 and data:
            report = data
    
    return render_template('accounting/balance_sheet.html', 
                          title='Statement of Financial Position', 
                          report=report,
                          as_of_date=as_of_date)


@bp.route('/profit-loss')
@login_required
@permission_required('reports:view')
def profit_loss():
    """Profit & Loss Report"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    report = {
        'revenue': [],
        'expenses': [],
        'total_revenue': 0,
        'total_expenses': 0,
        'net_income': 0,
        'start_date': start_date,
        'end_date': end_date
    }
    
    if start_date and end_date:
        api_url = f'/accounting/reports/income-statement?start_date={start_date}&end_date={end_date}'
        data, status = api_request('GET', api_url)
        
        if status == 200 and data:
            report = data
    
    return render_template('accounting/profit_loss.html', 
                          title='Profit & Loss', 
                          report=report,
                          start_date=start_date,
                          end_date=end_date)


@bp.route('/trial-balance')
@login_required
@permission_required('reports:view')
def trial_balance():
    """Trial Balance Report"""
    as_of_date = request.args.get('as_of_date', '')
    
    items = []
    total_debit = 0
    total_credit = 0
    
    if as_of_date:
        data, status = api_request('GET', f'/accounting/reports/trial-balance?as_of_date={as_of_date}')
        
        if status == 200 and data:
            items = data
            total_debit = sum(item.get('debit', 0) for item in items)
            total_credit = sum(item.get('credit', 0) for item in items)
    
    return render_template('accounting/trial_balance.html', 
                          title='Trial Balance', 
                          items=items,
                          total_debit=total_debit,
                          total_credit=total_credit,
                          as_of_date=as_of_date)


@bp.route('/trial-balance/export/pdf')
@login_required
@permission_required('reports:view')
def trial_balance_pdf():
    """Export Trial Balance to PDF"""
    as_of_date = request.args.get('as_of_date')
    
    items = []
    total_debit = 0
    total_credit = 0
    
    if as_of_date:
        data, status = api_request('GET', f'/accounting/reports/trial-balance?as_of_date={as_of_date}')
        
        if status == 200 and data:
            items = data
            total_debit = sum(item.get('debit', 0) for item in items)
            total_credit = sum(item.get('credit', 0) for item in items)
    
    business = {
        'name': session.get('business_name', 'Company'),
        'branch': session.get('selected_branch_name', 'Main Branch')
    }
    
    report = {
        'accounts': items,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'is_balanced': abs(total_debit - total_credit) < 0.01
    }
    
    html = render_template('reports/pdf/trial_balance_pdf.html',
                          report=report,
                          business=business,
                          as_of_date=as_of_date,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=trial_balance_{as_of_date or date.today()}.pdf'
    return response


# ==================== BALANCE SHEET EXPORTS ====================

@bp.route('/balance-sheet/export/pdf')
@login_required
@permission_required('reports:view')
def balance_sheet_pdf():
    """Export Balance Sheet to PDF"""
    as_of_date = request.args.get('as_of_date')
    
    report = {
        'non_current_assets': {'fixed_assets_cost': 0, 'accumulated_depreciation': 0, 'net_book_value': 0, 'other_non_current': 0, 'total': 0},
        'current_assets': {'inventory': 0, 'accounts_receivable': 0, 'vat_receivable': 0, 'cash_and_bank': 0, 'other_current_assets': 0, 'vendor_advances': 0, 'total': 0},
        'total_assets': 0,
        'liabilities': {'accounts_payable': 0, 'payroll_liabilities': 0, 'paye_payable': 0, 'pension_payable': 0, 'vat_payable': 0, 'customer_advances': 0, 'other_liabilities': 0, 'total': 0},
        'equity': {'owners_equity': 0, 'retained_earnings': 0, 'opening_balance_equity': 0, 'current_period_earnings': 0, 'total': 0},
        'total_liabilities': 0,
        'total_equity': 0,
        'total_equity_and_liabilities': 0
    }
    
    if as_of_date:
        data, status = api_request('GET', f'/accounting/reports/balance-sheet?as_of_date={as_of_date}')
        if status == 200 and data:
            report = data
    
    business = {
        'name': session.get('business_name', 'Company'),
        'branch': session.get('selected_branch_name', 'Main Branch')
    }
    
    html = render_template('reports/pdf/balance_sheet_pdf.html',
                          report=report,
                          business=business,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=statement_of_financial_position_{as_of_date or date.today()}.pdf'
    return response


@bp.route('/balance-sheet/export/excel')
@login_required
@permission_required('reports:view')
def balance_sheet_excel():
    """Export Balance Sheet to Excel"""
    as_of_date = request.args.get('as_of_date')
    
    report = {
        'non_current_assets': {'fixed_assets_cost': 0, 'accumulated_depreciation': 0, 'net_book_value': 0, 'other_non_current': 0, 'total': 0},
        'current_assets': {'inventory': 0, 'accounts_receivable': 0, 'vat_receivable': 0, 'cash_and_bank': 0, 'other_current_assets': 0, 'vendor_advances': 0, 'total': 0},
        'total_assets': 0,
        'liabilities': {'accounts_payable': 0, 'payroll_liabilities': 0, 'paye_payable': 0, 'pension_payable': 0, 'vat_payable': 0, 'customer_advances': 0, 'other_liabilities': 0, 'total': 0},
        'equity': {'owners_equity': 0, 'retained_earnings': 0, 'opening_balance_equity': 0, 'current_period_earnings': 0, 'total': 0},
        'total_liabilities': 0,
        'total_equity': 0,
        'total_equity_and_liabilities': 0
    }
    
    if as_of_date:
        data, status = api_request('GET', f'/accounting/reports/balance-sheet?as_of_date={as_of_date}')
        if status == 200 and data:
            report = data
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement of Financial Position"
    
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11, color="333333")
    subsection_font = Font(bold=True, size=10, italic=True, color="666666")
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    asset_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    equity_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    subtotal_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Statement of Financial Position"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"As of: {as_of_date or date.today().isoformat()}"
    ws.merge_cells('A2:D2')
    
    row = 4
    
    # Column headers
    headers = ['Description', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = thin_border
    row += 2
    
    # ASSETS SECTION
    ws.cell(row=row, column=1, value="ASSETS").font = header_font
    ws.cell(row=row, column=1).fill = asset_fill
    ws.cell(row=row, column=2).fill = asset_fill
    row += 1
    
    # Non-Current Assets
    ws.cell(row=row, column=1, value="Non-Current Assets").font = subsection_font
    row += 1
    
    ws.cell(row=row, column=1, value="Fixed Assets (Cost)").border = thin_border
    ws.cell(row=row, column=2, value=report['non_current_assets']['fixed_assets_cost']).number_format = currency_format
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    ws.cell(row=row, column=1, value="Less: Accumulated Depreciation").border = thin_border
    ws.cell(row=row, column=2, value=abs(report['non_current_assets']['accumulated_depreciation'])).number_format = currency_format
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    ws.cell(row=row, column=1, value="Net Book Value").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=report['non_current_assets']['net_book_value']).number_format = currency_format
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    if report['non_current_assets']['other_non_current'] != 0:
        ws.cell(row=row, column=1, value="Other Non-Current Assets").border = thin_border
        ws.cell(row=row, column=2, value=report['non_current_assets']['other_non_current']).number_format = currency_format
        ws.cell(row=row, column=2).border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total Non-Current Assets").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=2, value=report['non_current_assets']['total']).number_format = currency_format
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = subtotal_fill
    row += 2
    
    # Current Assets
    ws.cell(row=row, column=1, value="Current Assets").font = subsection_font
    row += 1
    
    current_asset_items = [
        ('Inventory', report['current_assets']['inventory']),
        ('Accounts Receivable', report['current_assets']['accounts_receivable']),
        ('VAT Receivable (Input VAT)', report['current_assets']['vat_receivable']),
        ('Cash & Bank', report['current_assets']['cash_and_bank']),
    ]
    
    if report['current_assets']['vendor_advances'] != 0:
        current_asset_items.append(('Vendor Advances', report['current_assets']['vendor_advances']))
    if report['current_assets']['other_current_assets'] != 0:
        current_asset_items.append(('Other Current Assets', report['current_assets']['other_current_assets']))
    
    for item_name, item_value in current_asset_items:
        ws.cell(row=row, column=1, value=item_name).border = thin_border
        ws.cell(row=row, column=2, value=item_value).number_format = currency_format
        ws.cell(row=row, column=2).border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total Current Assets").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=2, value=report['current_assets']['total']).number_format = currency_format
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = subtotal_fill
    row += 2
    
    # Total Assets
    ws.cell(row=row, column=1, value="TOTAL ASSETS").font = header_font
    ws.cell(row=row, column=1).fill = asset_fill
    ws.cell(row=row, column=2, value=report['total_assets']).number_format = currency_format
    ws.cell(row=row, column=2).font = header_font
    ws.cell(row=row, column=2).fill = asset_fill
    row += 3
    
    # EQUITY & LIABILITIES SECTION
    ws.cell(row=row, column=1, value="EQUITY & LIABILITIES").font = header_font
    ws.cell(row=row, column=1).fill = equity_fill
    ws.cell(row=row, column=2).fill = equity_fill
    row += 1
    
    # Liabilities
    ws.cell(row=row, column=1, value="Liabilities").font = subsection_font
    row += 1
    
    liability_items = [
        ('Accounts Payable', report['liabilities']['accounts_payable']),
        ('Payroll Liabilities', report['liabilities']['payroll_liabilities']),
        ('- PAYE Payable', report['liabilities']['paye_payable']),
        ('- Pension Payable', report['liabilities']['pension_payable']),
        ('VAT Payable (Output VAT)', report['liabilities']['vat_payable']),
    ]
    
    if report['liabilities']['customer_advances'] != 0:
        liability_items.append(('Customer Advances', report['liabilities']['customer_advances']))
    if report['liabilities']['other_liabilities'] != 0:
        liability_items.append(('Other Liabilities', report['liabilities']['other_liabilities']))
    
    for item_name, item_value in liability_items:
        ws.cell(row=row, column=1, value=item_name).border = thin_border
        ws.cell(row=row, column=2, value=item_value).number_format = currency_format
        ws.cell(row=row, column=2).border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total Liabilities").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=2, value=report['liabilities']['total']).number_format = currency_format
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = subtotal_fill
    row += 2
    
    # Equity
    ws.cell(row=row, column=1, value="Equity").font = subsection_font
    row += 1
    
    equity_items = [
        ("Owner's Equity", report['equity']['owners_equity']),
        ('Retained Earnings (Current Period)', report['equity']['current_period_earnings']),
    ]
    
    if report['equity']['retained_earnings'] != 0:
        equity_items.append(('Retained Earnings (Prior Periods)', report['equity']['retained_earnings']))
    if report['equity']['opening_balance_equity'] != 0:
        equity_items.append(('Opening Balance Equity', report['equity']['opening_balance_equity']))
    
    for item_name, item_value in equity_items:
        ws.cell(row=row, column=1, value=item_name).border = thin_border
        ws.cell(row=row, column=2, value=item_value).number_format = currency_format
        ws.cell(row=row, column=2).border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total Equity").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=2, value=report['equity']['total']).number_format = currency_format
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = subtotal_fill
    row += 2
    
    # Total Equity & Liabilities
    ws.cell(row=row, column=1, value="TOTAL EQUITY & LIABILITIES").font = header_font
    ws.cell(row=row, column=1).fill = equity_fill
    ws.cell(row=row, column=2, value=report['total_equity_and_liabilities']).number_format = currency_format
    ws.cell(row=row, column=2).font = header_font
    ws.cell(row=row, column=2).fill = equity_fill
    row += 2
    
    # Balance Check
    diff = abs(report['total_assets'] - report['total_equity_and_liabilities'])
    if diff > 0.01:
        ws.cell(row=row, column=1, value=f"Warning: Balance Sheet difference of {diff:.2f}")
        ws.cell(row=row, column=1).font = Font(color="FF0000")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=statement_of_financial_position_{as_of_date or date.today()}.xlsx'
    return response


# ==================== PROFIT & LOSS EXPORTS ====================

@bp.route('/profit-loss/export/pdf')
@login_required
@permission_required('reports:view')
def profit_loss_pdf():
    """Export Profit & Loss to PDF"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    report = {
        'revenue': [], 'expenses': [],
        'total_revenue': 0, 'total_expenses': 0, 'net_income': 0,
        'start_date': start_date, 'end_date': end_date
    }
    
    if start_date and end_date:
        data, status = api_request('GET', f'/accounting/reports/income-statement?start_date={start_date}&end_date={end_date}')
        if status == 200 and data:
            report = data
    
    business = {
        'name': session.get('business_name', 'Company'),
        'branch': session.get('selected_branch_name', 'Main Branch')
    }
    
    html = render_template('reports/pdf/profit_loss_pdf.html',
                          report=report,
                          business=business,
                          now=datetime.now())
    
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=profit_loss_{end_date or date.today()}.pdf'
    return response


@bp.route('/profit-loss/export/excel')
@login_required
@permission_required('reports:view')
def profit_loss_excel():
    """Export Profit & Loss to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    report = {
        'revenue': [], 'expenses': [],
        'total_revenue': 0, 'total_expenses': 0, 'net_income': 0,
        'start_date': start_date, 'end_date': end_date
    }
    
    if start_date and end_date:
        data, status = api_request('GET', f'/accounting/reports/income-statement?start_date={start_date}&end_date={end_date}')
        if status == 200 and data:
            report = data
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    currency_format = '#,##0.00'
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    profit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Profit & Loss Statement"
    ws['A1'].font = header_font
    ws.merge_cells('A1:C1')
    
    ws['A2'] = f"Period: {start_date or 'N/A'} to {end_date or 'N/A'}"
    ws.merge_cells('A2:C2')
    
    row = 4
    
    # Revenue Section
    ws.cell(row=row, column=1, value="REVENUE").font = section_font
    ws.cell(row=row, column=1).font = Font(bold=True, color="059669")
    row += 1
    
    headers = ['Code', 'Account', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    for item in report.get('revenue', []):
        ws.cell(row=row, column=1, value=item.get('account_code')).border = thin_border
        ws.cell(row=row, column=2, value=item.get('account_name')).border = thin_border
        amount_cell = ws.cell(row=row, column=3, value=item.get('balance'))
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
        row += 1
    
    # Total Revenue
    total_cell = ws.cell(row=row, column=2, value="Total Revenue")
    total_cell.font = Font(bold=True)
    total_cell.fill = total_fill
    ws.cell(row=row, column=3, value=report.get('total_revenue', 0)).number_format = currency_format
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=3).fill = total_fill
    row += 2
    
    # Expenses Section
    ws.cell(row=row, column=1, value="EXPENSES").font = section_font
    ws.cell(row=row, column=1).font = Font(bold=True, color="DC2626")
    row += 1
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    for item in report.get('expenses', []):
        ws.cell(row=row, column=1, value=item.get('account_code')).border = thin_border
        ws.cell(row=row, column=2, value=item.get('account_name')).border = thin_border
        amount_cell = ws.cell(row=row, column=3, value=item.get('balance'))
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
        row += 1
    
    # Total Expenses
    total_cell = ws.cell(row=row, column=2, value="Total Expenses")
    total_cell.font = Font(bold=True)
    total_cell.fill = total_fill
    ws.cell(row=row, column=3, value=report.get('total_expenses', 0)).number_format = currency_format
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=3).fill = total_fill
    row += 2
    
    # Net Income
    net_income = report.get('net_income', 0)
    ws.cell(row=row, column=2, value="NET INCOME (PROFIT/LOSS)").font = Font(bold=True, size=12)
    net_cell = ws.cell(row=row, column=3, value=net_income)
    net_cell.number_format = currency_format
    net_cell.font = Font(bold=True, size=12)
    if net_income >= 0:
        net_cell.fill = profit_fill
    else:
        net_cell.fill = loss_fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=profit_loss_{end_date or date.today()}.xlsx'
    return response
