"""
Banking API Routes - Bank Accounts, Fund Transfers, Reconciliation
"""
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List
from datetime import date
from decimal import Decimal

from app.core.database import get_db
from app.core.security import get_current_active_user, PermissionChecker
from app.schemas import (
    BankAccountCreate, BankAccountResponse,
    FundTransferCreate, FundTransferResponse,
    DepositRequest, WithdrawRequest, ReconcileRequest
)
from app.services.banking_service import BankAccountService, FundTransferService

router = APIRouter(prefix="/banking", tags=["Banking"])


# ==================== PAYMENT ACCOUNTS ====================

@router.get("/payment-accounts")
async def list_payment_accounts(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """List all accounts available for payments (bank accounts + cash accounts)"""
    from app.models import Account, CashBookEntry
    from sqlalchemy import func

    # Get all bank accounts for this branch
    bank_account_service = BankAccountService(db)
    bank_accounts = bank_account_service.get_by_branch(
        current_user.selected_branch.id,
        current_user.business_id
    )

    # Also get Cash-type accounts from Chart of Accounts
    cash_accounts = db.query(Account).filter(
        Account.business_id == current_user.business_id,
        Account.name.ilike('%cash%'),
        Account.is_active == True
    ).all()

    def calculate_cashbook_balance(account_id):
        """Helper to calculate balance from CashBookEntry"""
        receipts = db.query(func.sum(CashBookEntry.amount)).filter(
            CashBookEntry.account_id == account_id,
            CashBookEntry.branch_id == current_user.selected_branch.id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.entry_type == 'receipt'
        ).scalar() or 0

        payments = db.query(func.sum(CashBookEntry.amount)).filter(
            CashBookEntry.account_id == account_id,
            CashBookEntry.branch_id == current_user.selected_branch.id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.entry_type == 'payment'
        ).scalar() or 0

        transfers_in = db.query(func.sum(CashBookEntry.amount)).filter(
            CashBookEntry.account_id == account_id,
            CashBookEntry.branch_id == current_user.selected_branch.id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.is_transfer == True,
            CashBookEntry.transfer_direction == 'in'
        ).scalar() or 0

        transfers_out = db.query(func.sum(CashBookEntry.amount)).filter(
            CashBookEntry.account_id == account_id,
            CashBookEntry.branch_id == current_user.selected_branch.id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.is_transfer == True,
            CashBookEntry.transfer_direction == 'out'
        ).scalar() or 0

        return float(receipts) + float(transfers_in) - float(payments) - float(transfers_out)

    # Build combined list
    result = []

    # Add bank accounts
    for ba in bank_accounts:
        balance = calculate_cashbook_balance(ba.chart_of_account_id) if ba.chart_of_account_id else 0

        result.append({
            'id': ba.chart_of_account_id if ba.chart_of_account_id else ba.id,
            'bank_account_id': ba.id,
            'type': 'bank',
            'name': ba.account_name,
            'bank_name': ba.bank_name,
            'account_number': ba.account_number,
            'balance': balance,
            'currency': ba.currency,
            'is_active': ba.is_active if hasattr(ba, 'is_active') else True
        })
    
    # Add cash accounts from COA
    for ca in cash_accounts:
        # Skip if this account is already linked to a bank account
        if ca.bank_accounts:
            continue
        
        balance = calculate_cashbook_balance(ca.id)
        
        result.append({
            'id': ca.id,
            'type': 'cash',
            'name': ca.name,
            'bank_name': None,
            'account_number': ca.code,
            'balance': balance,
            'currency': 'USD',
            'is_active': ca.is_active
        })
    
    # If no accounts found, add default Cash account
    if not result:
        # Try to find or create default cash account
        default_cash = db.query(Account).filter(
            Account.business_id == current_user.business_id,
            Account.name == 'Cash'
        ).first()
        
        if default_cash:
            balance = calculate_cashbook_balance(default_cash.id)
            
            result.append({
                'id': default_cash.id,
                'type': 'cash',
                'name': 'Cash',
                'bank_name': None,
                'account_number': default_cash.code,
                'balance': balance,
                'currency': 'USD',
                'is_active': True
            })
    
    return result


# ==================== BANK ACCOUNTS ====================

@router.get("/accounts")
async def list_bank_accounts(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """List all bank accounts with calculated balance from CashBookEntry"""
    from app.models import CashBookEntry
    from sqlalchemy import func

    account_service = BankAccountService(db)
    bank_accounts = account_service.get_by_branch(
        current_user.selected_branch.id,
        current_user.business_id
    )

    result = []
    for ba in bank_accounts:
        # Calculate balance from CashBookEntry for comprehensive tracking
        receipts = db.query(func.sum(CashBookEntry.amount)).filter(
            CashBookEntry.account_id == ba.chart_of_account_id,
            CashBookEntry.branch_id == current_user.selected_branch.id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.entry_type == 'receipt'
        ).scalar() or 0

        payments = db.query(func.sum(CashBookEntry.amount)).filter(
            CashBookEntry.account_id == ba.chart_of_account_id,
            CashBookEntry.branch_id == current_user.selected_branch.id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.entry_type == 'payment'
        ).scalar() or 0

        transfers_in = db.query(func.sum(CashBookEntry.amount)).filter(
            CashBookEntry.account_id == ba.chart_of_account_id,
            CashBookEntry.branch_id == current_user.selected_branch.id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.is_transfer == True,
            CashBookEntry.transfer_direction == 'in'
        ).scalar() or 0

        transfers_out = db.query(func.sum(CashBookEntry.amount)).filter(
            CashBookEntry.account_id == ba.chart_of_account_id,
            CashBookEntry.branch_id == current_user.selected_branch.id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.is_transfer == True,
            CashBookEntry.transfer_direction == 'out'
        ).scalar() or 0

        balance = float(receipts) + float(transfers_in) - float(payments) - float(transfers_out)

        result.append({
            'id': ba.id,
            'account_name': ba.account_name,
            'bank_name': ba.bank_name,
            'account_number': ba.account_number,
            'currency': ba.currency,
            'opening_balance': float(ba.opening_balance) if ba.opening_balance else 0.0,
            'current_balance': balance,  # Calculated from CashBookEntry
            'last_reconciliation_date': ba.last_reconciliation_date.isoformat() if ba.last_reconciliation_date else None,
            'last_reconciliation_balance': float(ba.last_reconciliation_balance) if ba.last_reconciliation_balance else None,
            'chart_of_account_id': ba.chart_of_account_id,
            'chart_of_account_name': ba.chart_of_account.name if ba.chart_of_account else None,
            'branch_id': ba.branch_id,
            'business_id': ba.business_id
        })

    return result


@router.post("/accounts", response_model=BankAccountResponse, dependencies=[Depends(PermissionChecker(["bank:create"]))])
async def create_bank_account(
    account_data: BankAccountCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Create a new bank account"""
    from app.models import BankAccount, Account

    # If no chart_of_account_id provided, create a new Bank account in COA
    if not account_data.chart_of_account_id:
        # Generate a unique code for the new account
        last_bank_account = db.query(Account).filter(
            Account.business_id == current_user.business_id,
            Account.type == "Asset"
        ).order_by(Account.id.desc()).first()

        if last_bank_account and last_bank_account.code:
            try:
                # Try to increment the code
                code_num = int(last_bank_account.code.replace("1", "").lstrip("0") or "0") + 1
                new_code = f"1{code_num:03d}"
            except ValueError:
                new_code = "1101"
        else:
            new_code = "1101"

        # Create new Bank account in Chart of Accounts
        new_coa_account = Account(
            name=f"Bank - {account_data.account_name}",
            code=new_code,
            type="Asset",
            description=f"Bank account: {account_data.bank_name or account_data.account_name}",
            business_id=current_user.business_id
        )
        db.add(new_coa_account)
        db.flush()
        account_data.chart_of_account_id = new_coa_account.id
    else:
        # Check if chart_of_account_id is already linked to another bank account
        existing_bank = db.query(BankAccount).filter(
            BankAccount.chart_of_account_id == account_data.chart_of_account_id,
            BankAccount.business_id == current_user.business_id
        ).first()
        if existing_bank:
            raise HTTPException(
                status_code=400,
                detail=f"This account is already linked to bank account '{existing_bank.account_name}'. Please select a different account or leave blank to auto-create one."
            )

    account_service = BankAccountService(db)
    account = account_service.create(
        account_data,
        current_user.selected_branch.id,
        current_user.business_id
    )
    db.commit()
    return account


@router.get("/accounts/{account_id}")
async def get_bank_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Get bank account by ID with calculated balance from CashBookEntry"""
    from app.models import CashBookEntry
    from sqlalchemy import func

    account_service = BankAccountService(db)
    account = account_service.get_by_id(account_id, current_user.business_id)
    if not account:
        raise HTTPException(status_code=404, detail="Bank account not found")

    # Calculate balance from CashBookEntry for comprehensive tracking
    # Receipts (money in) increase balance, Payments (money out) decrease balance
    receipts = db.query(func.sum(CashBookEntry.amount)).filter(
        CashBookEntry.account_id == account.chart_of_account_id,
        CashBookEntry.branch_id == current_user.selected_branch.id,
        CashBookEntry.business_id == current_user.business_id,
        CashBookEntry.entry_type == 'receipt'
    ).scalar() or 0

    payments = db.query(func.sum(CashBookEntry.amount)).filter(
        CashBookEntry.account_id == account.chart_of_account_id,
        CashBookEntry.branch_id == current_user.selected_branch.id,
        CashBookEntry.business_id == current_user.business_id,
        CashBookEntry.entry_type == 'payment'
    ).scalar() or 0

    # Also account for transfers - transfers in are like receipts, transfers out are like payments
    transfers_in = db.query(func.sum(CashBookEntry.amount)).filter(
        CashBookEntry.account_id == account.chart_of_account_id,
        CashBookEntry.branch_id == current_user.selected_branch.id,
        CashBookEntry.business_id == current_user.business_id,
        CashBookEntry.is_transfer == True,
        CashBookEntry.transfer_direction == 'in'
    ).scalar() or 0

    transfers_out = db.query(func.sum(CashBookEntry.amount)).filter(
        CashBookEntry.account_id == account.chart_of_account_id,
        CashBookEntry.branch_id == current_user.selected_branch.id,
        CashBookEntry.business_id == current_user.business_id,
        CashBookEntry.is_transfer == True,
        CashBookEntry.transfer_direction == 'out'
    ).scalar() or 0

    # Calculate balance: receipts + transfers_in - payments - transfers_out
    balance = float(receipts) + float(transfers_in) - float(payments) - float(transfers_out)
    
    return {
        'id': account.id,
        'account_name': account.account_name,
        'bank_name': account.bank_name,
        'account_number': account.account_number,
        'currency': account.currency,
        'opening_balance': float(account.opening_balance) if account.opening_balance else 0.0,
        'current_balance': balance,  # Calculated from CashBookEntry
        'last_reconciliation_date': account.last_reconciliation_date.isoformat() if account.last_reconciliation_date else None,
        'last_reconciliation_balance': float(account.last_reconciliation_balance) if account.last_reconciliation_balance else None,
        'chart_of_account_id': account.chart_of_account_id,
        'chart_of_account_name': account.chart_of_account.name if account.chart_of_account else None,
        'branch_id': account.branch_id,
        'business_id': account.business_id
    }


@router.post("/accounts/{account_id}/deposit", dependencies=[Depends(PermissionChecker(["bank:create"]))])
async def deposit_to_account(
    account_id: int,
    deposit_data: DepositRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Deposit to bank account"""
    account_service = BankAccountService(db)
    try:
        account = account_service.deposit(
            account_id,
            current_user.business_id,
            deposit_data.amount,
            deposit_data.description
        )
        db.commit()
        return {"message": "Deposit successful", "balance": account.current_balance}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/accounts/{account_id}/withdraw", dependencies=[Depends(PermissionChecker(["bank:create"]))])
async def withdraw_from_account(
    account_id: int,
    withdraw_data: WithdrawRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Withdraw from bank account"""
    account_service = BankAccountService(db)
    try:
        account = account_service.withdraw(
            account_id,
            current_user.business_id,
            withdraw_data.amount,
            withdraw_data.description
        )
        db.commit()
        return {"message": "Withdrawal successful", "balance": account.current_balance}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/accounts/{account_id}/reconcile", dependencies=[Depends(PermissionChecker(["bank:reconcile"]))])
async def reconcile_account(
    account_id: int,
    reconcile_data: ReconcileRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Reconcile bank account"""
    account_service = BankAccountService(db)
    try:
        result = account_service.reconcile(
            account_id,
            current_user.business_id,
            reconcile_data.statement_balance,
            reconcile_data.reconciliation_date
        )
        db.commit()
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/accounts/{account_id}", dependencies=[Depends(PermissionChecker(["bank:create"]))])
async def delete_bank_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Delete bank account"""
    account_service = BankAccountService(db)
    if not account_service.delete(account_id, current_user.business_id):
        raise HTTPException(status_code=400, detail="Cannot delete account with transfers")
    db.commit()
    return {"message": "Bank account deleted"}


# ==================== FUND TRANSFERS ====================

@router.get("/transfers")
async def list_transfers(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """List all fund transfers"""
    transfer_service = FundTransferService(db)
    transfers = transfer_service.get_by_branch(
        current_user.selected_branch.id,
        current_user.business_id
    )
    
    result = []
    for t in transfers:
        result.append({
            'id': t.id,
            'transfer_number': t.transfer_number,
            'transfer_date': t.transfer_date.isoformat() if t.transfer_date else None,
            'amount': float(t.amount) if t.amount else 0.0,
            'description': t.description,
            'reference': t.reference,
            'from_account_id': t.from_account_id,
            'from_account_type': t.from_account_type,
            'from_account_name': t.from_account_name,
            'to_account_id': t.to_account_id,
            'to_account_type': t.to_account_type,
            'to_account_name': t.to_account_name,
            'branch_id': t.branch_id,
            'business_id': t.business_id,
            'created_at': t.created_at.isoformat() if t.created_at else None
        })
    
    return result


@router.post("/transfers", response_model=FundTransferResponse, dependencies=[Depends(PermissionChecker(["bank:create"]))])
async def create_transfer(
    transfer_data: FundTransferCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Create a fund transfer"""
    transfer_service = FundTransferService(db)
    try:
        transfer = transfer_service.create(
            transfer_data,
            current_user.selected_branch.id,
            current_user.business_id
        )
        db.commit()
        return transfer
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/transfers/{transfer_id}", response_model=FundTransferResponse)
async def get_transfer(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Get fund transfer by ID"""
    transfer_service = FundTransferService(db)
    transfer = transfer_service.get_by_id(transfer_id, current_user.business_id)
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer not found")
    return transfer


@router.get("/accounts/{account_id}/transfer-history")
async def get_transfer_history(
    account_id: int,
    start_date: date = None,
    end_date: date = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Get transfer history for an account"""
    transfer_service = FundTransferService(db)
    return transfer_service.get_transfer_history(
        account_id,
        current_user.business_id,
        start_date,
        end_date
    )


@router.get("/accounts/{account_id}/transactions")
async def get_account_transactions(
    account_id: int,
    start_date: date = None,
    end_date: date = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Get all transactions for a bank account using CashBookEntry for comprehensive tracking"""
    from app.models import CashBookEntry, BankAccount
    from sqlalchemy import func

    # First verify the bank account belongs to this user's business
    account_service = BankAccountService(db)
    bank_account = account_service.get_by_id(account_id, current_user.business_id)
    if not bank_account:
        raise HTTPException(status_code=404, detail="Bank account not found")

    branch_id = current_user.selected_branch.id

    # Build query for CashBookEntry
    query = db.query(CashBookEntry).filter(
        CashBookEntry.branch_id == branch_id,
        CashBookEntry.business_id == current_user.business_id,
        CashBookEntry.account_id == bank_account.chart_of_account_id
    )

    # Apply date filters
    if start_date:
        query = query.filter(CashBookEntry.entry_date >= start_date)
    if end_date:
        query = query.filter(CashBookEntry.entry_date <= end_date)

    # Get total count for pagination
    total_count = query.count()

    # Get paginated entries ordered by date desc
    entries = query.order_by(
        CashBookEntry.entry_date.desc(),
        CashBookEntry.id.desc()
    ).offset(offset).limit(limit).all()

    # Get all entries for running balance calculation (only if no date filter)
    if not start_date and not end_date:
        all_entries = db.query(CashBookEntry).filter(
            CashBookEntry.branch_id == branch_id,
            CashBookEntry.business_id == current_user.business_id,
            CashBookEntry.account_id == bank_account.chart_of_account_id
        ).order_by(CashBookEntry.entry_date, CashBookEntry.id).all()
    else:
        # For filtered view, calculate balance within date range
        all_entries = []

    # Calculate running balance from all entries
    balance_map = {}
    current_balance = 0.0
    for entry in all_entries:
        if entry.entry_type == 'receipt' or (entry.is_transfer and entry.transfer_direction == 'in'):
            current_balance += float(entry.amount or 0)
        elif entry.entry_type == 'payment' or (entry.is_transfer and entry.transfer_direction == 'out'):
            current_balance -= float(entry.amount or 0)
        balance_map[entry.id] = current_balance

    # Format results
    result = []
    for entry in entries:
        # Determine debit/credit based on entry type
        if entry.entry_type == 'receipt' or (entry.is_transfer and entry.transfer_direction == 'in'):
            debit = float(entry.amount or 0)
            credit = 0.0
        elif entry.entry_type == 'payment' or (entry.is_transfer and entry.transfer_direction == 'out'):
            debit = 0.0
            credit = float(entry.amount or 0)
        else:
            # adjustment or other
            debit = float(entry.amount or 0) if entry.amount and float(entry.amount) > 0 else 0.0
            credit = abs(float(entry.amount or 0)) if entry.amount and float(entry.amount) < 0 else 0.0

        result.append({
            'id': entry.id,
            'date': entry.entry_date.isoformat() if entry.entry_date else None,
            'entry_number': entry.entry_number,
            'description': entry.description,
            'reference': entry.reference,
            'payee_payer': entry.payee_payer,
            'source_type': entry.source_type,
            'entry_type': entry.entry_type,
            'debit': debit,
            'credit': credit,
            'balance': balance_map.get(entry.id, float(entry.balance_after or 0)),
            'is_transfer': entry.is_transfer,
            'transfer_direction': entry.transfer_direction
        })

    return {
        'transactions': result,
        'total_count': total_count,
        'limit': limit,
        'offset': offset
    }


@router.get("/accounts/{account_id}/statement/pdf")
async def generate_bank_statement_pdf(
    account_id: int,
    start_date: date = None,
    end_date: date = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Generate PDF bank statement"""
    from app.models import CashBookEntry, Business, Branch
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from datetime import datetime

    # Verify bank account
    account_service = BankAccountService(db)
    bank_account = account_service.get_by_id(account_id, current_user.business_id)
    if not bank_account:
        raise HTTPException(status_code=404, detail="Bank account not found")

    # Get business and branch info
    business = db.query(Business).filter(Business.id == current_user.business_id).first()
    branch = db.query(Branch).filter(Branch.id == current_user.selected_branch.id).first()

    # Get transactions
    query = db.query(CashBookEntry).filter(
        CashBookEntry.branch_id == current_user.selected_branch.id,
        CashBookEntry.business_id == current_user.business_id,
        CashBookEntry.account_id == bank_account.chart_of_account_id
    )

    if start_date:
        query = query.filter(CashBookEntry.entry_date >= start_date)
    if end_date:
        query = query.filter(CashBookEntry.entry_date <= end_date)

    entries = query.order_by(CashBookEntry.entry_date, CashBookEntry.id).all()

    # Calculate balances
    running_balance = 0.0
    transactions = []
    for entry in entries:
        if entry.entry_type == 'receipt' or (entry.is_transfer and entry.transfer_direction == 'in'):
            debit = float(entry.amount or 0)
            credit = 0.0
            running_balance += debit
        else:
            debit = 0.0
            credit = float(entry.amount or 0)
            running_balance -= credit

        transactions.append({
            'date': entry.entry_date,
            'description': entry.description or entry.entry_number,
            'reference': entry.reference or '',
            'debit': debit,
            'credit': credit,
            'balance': running_balance
        })

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           leftMargin=0.5*inch, rightMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=9)

    elements = []

    # Header
    elements.append(Paragraph(business.name if business else "Bank Statement", title_style))
    elements.append(Paragraph(branch.name if branch else "", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))

    # Account Info
    elements.append(Paragraph(f"<b>Account:</b> {bank_account.account_name}", styles['Normal']))
    elements.append(Paragraph(f"<b>Bank:</b> {bank_account.bank_name or '-'}", styles['Normal']))
    elements.append(Paragraph(f"<b>Account No:</b> {bank_account.account_number or '-'}", styles['Normal']))

    # Date range
    date_range = f"Period: {start_date or 'All'} to {end_date or 'Present'}"
    elements.append(Paragraph(date_range, styles['Normal']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", header_style))
    elements.append(Spacer(1, 0.3*inch))

    # Table
    table_data = [['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Balance']]
    for t in transactions:
        table_data.append([
            t['date'].strftime('%Y-%m-%d') if t['date'] else '',
            t['description'][:40] if t['description'] else '',
            t['reference'][:15] if t['reference'] else '',
            f"{t['debit']:,.2f}" if t['debit'] > 0 else '',
            f"{t['credit']:,.2f}" if t['credit'] > 0 else '',
            f"{t['balance']:,.2f}"
        ])

    # Add totals row
    total_debit = sum(t['debit'] for t in transactions)
    total_credit = sum(t['credit'] for t in transactions)
    table_data.append(['', 'TOTAL', '', f"{total_debit:,.2f}", f"{total_credit:,.2f}", f"{running_balance:,.2f}"])

    table = Table(table_data, colWidths=[0.8*inch, 2.5*inch, 0.8*inch, 0.9*inch, 0.9*inch, 0.9*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')]),
    ]))

    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"bank_statement_{bank_account.account_name}_{datetime.now().strftime('%Y%m%d')}.pdf"

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/accounts/{account_id}/statement/excel")
async def generate_bank_statement_excel(
    account_id: int,
    start_date: date = None,
    end_date: date = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Generate Excel bank statement"""
    from app.models import CashBookEntry, Business, Branch
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from datetime import datetime

    # Verify bank account
    account_service = BankAccountService(db)
    bank_account = account_service.get_by_id(account_id, current_user.business_id)
    if not bank_account:
        raise HTTPException(status_code=404, detail="Bank account not found")

    # Get business and branch info
    business = db.query(Business).filter(Business.id == current_user.business_id).first()
    branch = db.query(Branch).filter(Branch.id == current_user.selected_branch.id).first()

    # Get transactions
    query = db.query(CashBookEntry).filter(
        CashBookEntry.branch_id == current_user.selected_branch.id,
        CashBookEntry.business_id == current_user.business_id,
        CashBookEntry.account_id == bank_account.chart_of_account_id
    )

    if start_date:
        query = query.filter(CashBookEntry.entry_date >= start_date)
    if end_date:
        query = query.filter(CashBookEntry.entry_date <= end_date)

    entries = query.order_by(CashBookEntry.entry_date, CashBookEntry.id).all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=10)
    currency_font = Font(name='Consolas', size=10)
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    ws['A1'] = business.name if business else "Bank Statement"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')

    ws['A2'] = branch.name if branch else ""
    ws['A2'].font = subtitle_font
    ws.merge_cells('A2:F2')

    # Account info
    ws['A4'] = "Account Name:"
    ws['B4'] = bank_account.account_name
    ws['A5'] = "Bank Name:"
    ws['B5'] = bank_account.bank_name or '-'
    ws['A6'] = "Account Number:"
    ws['B6'] = bank_account.account_number or '-'
    ws['A7'] = "Period:"
    ws['B7'] = f"{start_date or 'All'} to {end_date or 'Present'}"
    ws['A8'] = "Generated:"
    ws['B8'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    # Table headers
    headers = ['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    row = 11
    running_balance = 0.0
    total_debit = 0.0
    total_credit = 0.0

    for entry in entries:
        if entry.entry_type == 'receipt' or (entry.is_transfer and entry.transfer_direction == 'in'):
            debit = float(entry.amount or 0)
            credit = 0.0
            running_balance += debit
            total_debit += debit
        else:
            debit = 0.0
            credit = float(entry.amount or 0)
            running_balance -= credit
            total_credit += credit

        ws.cell(row=row, column=1, value=entry.entry_date.strftime('%Y-%m-%d') if entry.entry_date else '').border = thin_border
        ws.cell(row=row, column=2, value=entry.description or entry.entry_number).border = thin_border
        ws.cell(row=row, column=3, value=entry.reference or '').border = thin_border

        debit_cell = ws.cell(row=row, column=4, value=debit if debit > 0 else None)
        debit_cell.number_format = '#,##0.00'
        debit_cell.font = currency_font
        debit_cell.alignment = Alignment(horizontal='right')
        debit_cell.border = thin_border

        credit_cell = ws.cell(row=row, column=5, value=credit if credit > 0 else None)
        credit_cell.number_format = '#,##0.00'
        credit_cell.font = currency_font
        credit_cell.alignment = Alignment(horizontal='right')
        credit_cell.border = thin_border

        balance_cell = ws.cell(row=row, column=6, value=running_balance)
        balance_cell.number_format = '#,##0.00'
        balance_cell.font = currency_font
        balance_cell.alignment = Alignment(horizontal='right')
        balance_cell.border = thin_border

        row += 1

    # Totals row
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = total_font
        ws.cell(row=row, column=col).border = thin_border

    ws.cell(row=row, column=2, value="TOTAL")
    ws.cell(row=row, column=4, value=total_debit).number_format = '#,##0.00'
    ws.cell(row=row, column=5, value=total_credit).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=running_balance).number_format = '#,##0.00'

    # Adjust column widths
    column_widths = [12, 40, 15, 12, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"bank_statement_{bank_account.account_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
